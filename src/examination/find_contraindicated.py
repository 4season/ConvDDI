"""
find_contraindicated.py
=======================
데이터셋(TL_1~TL_8) 라벨 JSON에서 약물 정보를 추출하고,
병용금기 DB(contraindicated_drugs.xlsx)와 교차 검색합니다.

매칭 전략: Excel INN명이 JSON 풀 화학명의 substring인지 확인 (대소문자 무관)
예) Excel "rosuvastatin"  ←→  JSON "Rosuvastatin Calcium"  → 매칭
    Excel "metformin"    ←→  JSON "Metformin Hydrochloride" → 매칭
"""

import json
import glob
import os
import re
import openpyxl
import pandas as pd
from pathlib import Path
from collections import defaultdict


# ─────────────────────────── 경로 설정 ────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
LABEL_BASE   = PROJECT_ROOT / "data" / "labels"
EXCEL_PATH   = PROJECT_ROOT / "data" / "contraindicated_drugs.xlsx"


# ─────────────────────── 헬퍼 함수 ───────────────────────────────

def clean_excel_ingredient(raw: str) -> str:
    """Excel 성분명 정규화: 괄호 제거 → 소문자 → 공백 정리"""
    if not raw:
        return ""
    s = re.sub(r"\(.*?\)", "", str(raw))   # 괄호 및 내부 제거
    s = s.strip().lower()
    return s


def extract_ingredients(material_en: str) -> list[str]:
    """
    JSON dl_material_en 필드에서 개별 성분 목록 추출.
    '|' 구분자로 복합 성분을 분리하고 각 성분을 소문자 정규화.
    """
    if not material_en:
        return []
    parts = [p.strip().lower() for p in material_en.split("|") if p.strip()]
    return parts


def ingredient_matches(excel_inn: str, drug_ingredients: list[str]) -> bool:
    """
    Excel INN명이 약물의 임의 성분 풀 화학명에 substring으로 포함되는지 확인.
    단, 너무 짧은 INN(≤3자)은 오매칭 방지를 위해 정확 일치만 허용.
    """
    if not excel_inn:
        return False
    for ing in drug_ingredients:
        if len(excel_inn) <= 3:
            if excel_inn == ing:
                return True
        else:
            if excel_inn in ing:
                return True
    return False


# ────────────────────── Step 1: JSON 스캔 ─────────────────────────

def scan_dataset_drugs() -> dict:
    """
    TL_1~TL_8 전체 JSON을 순회하며 고유 약물 정보를 수집.
    반환: {drug_code: {'name': ..., 'material_en': ..., 'material_kr': ..., 'ingredients': [...]}}
    """
    pattern = str(LABEL_BASE / "**" / "*.json")
    json_files = glob.glob(pattern, recursive=True)
    print(f"[스캔] JSON 파일 수: {len(json_files):,}개")

    drugs: dict = {}
    errors = 0

    for fp in json_files:
        try:
            with open(fp, encoding="utf-8") as f:
                data = json.load(f)
            for img in data.get("images", []):
                code = img.get("dl_mapping_code")
                if code and code not in drugs:
                    material_en = img.get("dl_material_en", "") or ""
                    drugs[code] = {
                        "name":        img.get("dl_name", ""),
                        "material_en": material_en,
                        "material_kr": img.get("dl_material", ""),
                        "ingredients": extract_ingredients(material_en),
                    }
        except Exception:
            errors += 1

    print(f"[스캔] 고유 약물: {len(drugs)}개 | 오류: {errors}개")
    return drugs


# ────────────────────── Step 2: Excel 로드 ───────────────────────

def load_contraindicated_db() -> list[dict]:
    """
    병용금기 Excel을 파싱하여 정규화된 쌍 목록 반환.
    [{'m1': ..., 'm2': ..., 'note': ..., 'info': ...}, ...]
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    pairs = []
    for r in range(2, ws.max_row + 1):
        m1   = clean_excel_ingredient(ws.cell(r, 2).value)
        m2   = clean_excel_ingredient(ws.cell(r, 3).value)
        note = ws.cell(r, 4).value or ""
        info = ws.cell(r, 5).value or ""
        if m1 and m2:
            pairs.append({"m1": m1, "m2": m2, "note": note, "info": info})

    print(f"[Excel] 병용금기 쌍: {len(pairs):,}개 로드 완료")
    return pairs


# ────────────────────── Step 3: 교차 매칭 ────────────────────────

def find_contraindicated_pairs(drugs: dict, db: list[dict]) -> list[dict]:
    """
    데이터셋 약물과 병용금기 DB를 교차 검색.
    두 성분이 모두 데이터셋에 존재하는 쌍만 반환.
    """
    results = []

    for pair in db:
        m1, m2 = pair["m1"], pair["m2"]

        # 각 Excel 성분에 매칭되는 데이터셋 약물 수집
        matched_m1 = [
            code for code, info in drugs.items()
            if ingredient_matches(m1, info["ingredients"])
        ]
        matched_m2 = [
            code for code, info in drugs.items()
            if ingredient_matches(m2, info["ingredients"])
        ]

        if matched_m1 and matched_m2:
            # 같은 약이 m1=m2 매칭되는 경우(복합제) 제외
            if set(matched_m1) == set(matched_m2):
                continue
            results.append({
                "excel_m1":    m1,
                "excel_m2":    m2,
                "dataset_m1":  matched_m1,
                "dataset_m2":  matched_m2,
                "note":        pair["note"],
                "info":        pair["info"],
            })

    return results


# ─────────────────────── Step 4: 결과 출력 ───────────────────────

def print_report(drugs: dict, results: list[dict]) -> None:
    print("\n" + "=" * 70)
    print(f"  데이터셋 내 병용금기 쌍: {len(results)}개 발견")
    print("=" * 70)

    for i, r in enumerate(results, 1):
        m1_names = [drugs[c]["name"] for c in r["dataset_m1"]]
        m2_names = [drugs[c]["name"] for c in r["dataset_m2"]]
        print(f"\n[{i:03d}] {r['excel_m1']}  ↔  {r['excel_m2']}")
        print(f"      ▸ 성분1 약물: {', '.join(m1_names)}")
        print(f"      ▸ 성분2 약물: {', '.join(m2_names)}")
        if r["info"]:
            print(f"      ▸ 이유: {r['info']}")
        if r["note"]:
            print(f"      ▸ 조건: {r['note']}")

    print("\n" + "=" * 70)


def save_results_csv(drugs: dict, results: list[dict]) -> Path:
    """결과를 CSV로 저장해 추후 분석에 활용"""
    rows = []
    for r in results:
        for c1 in r["dataset_m1"]:
            for c2 in r["dataset_m2"]:
                rows.append({
                    "drug_code_1":    c1,
                    "drug_name_1":    drugs[c1]["name"],
                    "material_en_1":  drugs[c1]["material_en"],
                    "drug_code_2":    c2,
                    "drug_name_2":    drugs[c2]["name"],
                    "material_en_2":  drugs[c2]["material_en"],
                    "excel_inn_1":    r["excel_m1"],
                    "excel_inn_2":    r["excel_m2"],
                    "contraind_info": r["info"],
                    "contraind_note": r["note"],
                })
    df = pd.DataFrame(rows)
    out_path = PROJECT_ROOT / "output" / "contraindicated_in_dataset.csv"
    out_path.parent.mkdir(exist_ok=True)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"\n결과 저장 완료: {out_path}")
    return out_path


# ─────────────────────────── 실행 ────────────────────────────────

def main():
    print("=" * 70)
    print("  병용금기 교차 검색 시작")
    print("=" * 70 + "\n")

    # 1. 데이터셋 약물 수집
    drugs = scan_dataset_drugs()

    # 2. 병용금기 DB 로드
    db = load_contraindicated_db()

    # 3. 교차 매칭
    print("\n[매칭] 교차 검색 중...")
    results = find_contraindicated_pairs(drugs, db)

    # 4. 결과 출력 및 저장
    print_report(drugs, results)
    save_results_csv(drugs, results)

    # 5. 요약 통계: 가장 많이 등장한 약물
    from collections import Counter
    drug_freq: Counter = Counter()
    for r in results:
        for c in r["dataset_m1"] + r["dataset_m2"]:
            drug_freq[c] += 1

    print("\n▶ 병용금기 쌍에 가장 많이 등장하는 데이터셋 약물 (상위 20개):")
    print(f"  {'코드':<12} {'약물명':<40} {'등장 횟수'}")
    print("  " + "-" * 60)
    for code, cnt in drug_freq.most_common(20):
        print(f"  {code:<12} {drugs[code]['name'][:38]:<40} {cnt}회")


if __name__ == "__main__":
    main()
