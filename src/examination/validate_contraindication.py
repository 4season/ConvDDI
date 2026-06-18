"""
validate_contraindication.py
============================
병용금기 '성분 매칭' 로직의 정확도를 정량 검증하고, 더 견고한 매칭으로 교체한다.

배경: 기존 매칭의 약점
----------------------
기존 ingredient_matches() 는 'Excel INN 이 약물 성분 문자열의 substring 인가' 로
판단한다. 이 방식은 약물명이 다른 약을 포함(substring)하는 순간 오경보를 낸다.
실제 데이터셋에서 발견된 오매칭(거짓 양성):
    INN 'ephedrine'  ⊂ 'pseudoephedrine hydrochloride'   → 다른 약인데 매칭
    INN 'omeprazole' ⊂ 'esomeprazole magnesium ...'      → 다른 약인데 매칭
    INN 'ibuprofen'  ⊂ 'dexibuprofen'                    → 다른 약인데 매칭
금기 검출에서 거짓 양성은 '없는 위험을 경고'하는 것이라 신뢰도를 떨어뜨린다.

개선: 성분(component) 단위 토큰 부분집합 매칭
--------------------------------------------
1) 성분 문자열을 '|' 로 component 단위로 나눈다.
2) 각 component 를 단어(token) 집합으로 만든다(소문자, 괄호·기호 제거).
3) INN 의 모든 단어가 '하나의' component 안에 토큰으로 존재할 때만 매칭.
   - 'ephedrine' 은 'pseudoephedrine' 토큰과 다른 단어이므로 매칭 안 됨(정상).
   - 'rosuvastatin' 은 component 'rosuvastatin calcium' 의 토큰이므로 매칭(정상).
4) INN 에 '+' 가 있으면 복합제 표기로 보고, 각 부분이 (서로 다른 component라도)
   매칭되어야 전체 매칭으로 인정한다.
   - 'calcium chloride' 같은 두 단어 INN 은 한 component 안에 두 단어가 모두 있어야
     하므로, calcium 과 chloride 가 서로 다른 성분에 흩어진 종합비타민에는
     오매칭되지 않는다(기존 토큰 풀링 방식의 거짓 양성 제거).

이 스크립트는 (a) 손으로 라벨링한 검증셋으로 기존/개선 매칭의
precision·recall·accuracy 를 비교하고, (b) 전체 데이터셋에 개선 매칭을 적용해
금기 쌍 수와 '제거된 거짓 양성'을 보고한다.

실행
----
    python src/examination/validate_contraindication.py
"""

import json
import re
from pathlib import Path

import openpyxl

SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
CLASS_MAP    = PROJECT_ROOT / "data" / "class_map.json"
EXCEL_PATH   = PROJECT_ROOT / "data" / "contraindicated_drugs.xlsx"
OUT_DIR      = PROJECT_ROOT / "output"


# ───────────────────────── 정규화 ────────────────────────────────

def clean(s: str) -> str:
    """괄호 내용 제거 → 소문자 → 영문/숫자/공백/+ 외 제거."""
    s = re.sub(r"\(.*?\)", " ", str(s))
    s = s.lower()
    s = re.sub(r"[^a-z0-9+ ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def component_token_sets(material_en: str) -> list[set]:
    """'|' 로 나눈 각 성분(component)을 토큰 집합으로 변환."""
    sets = []
    for part in str(material_en).split("|"):
        toks = set(clean(part).replace("+", " ").split())
        if toks:
            sets.append(toks)
    return sets


# ─────────────────────── 두 가지 매칭 ────────────────────────────

def match_substring(inn: str, material_en: str) -> bool:
    """[기존] INN 이 임의 성분 문자열의 substring 인가 (≤3자는 정확 일치)."""
    inn = clean(inn)
    comps = [clean(p) for p in str(material_en).split("|") if p.strip()]
    for ing in comps:
        if len(inn) <= 3:
            if inn == ing:
                return True
        elif inn and inn in ing:
            return True
    return False


def match_token(inn: str, material_en: str) -> bool:
    """[개선] 성분 단위 토큰 부분집합 매칭 + '+' 복합제 처리."""
    comp_sets = component_token_sets(material_en)
    if not comp_sets:
        return False
    # '+' 로 복합제 부분 분리
    parts = [clean(p) for p in clean(inn).split("+")]
    parts = [p for p in parts if p]
    if not parts:
        return False
    for part in parts:
        words = part.split()
        # 이 part 의 모든 단어가 '하나의' component 안에 있어야 함
        if not any(all(w in cs for w in words) for cs in comp_sets):
            return False
    return True


# ─────────────── 손으로 라벨링한 검증셋 (gold set) ─────────────────
# (INN, 약물 성분 material_en, 정답 매칭 여부) — 약학적으로 직접 확인한 라벨.
GOLD = [
    # ── 매칭되어야 하는 참(True) 케이스 ──
    ("rosuvastatin", "Rosuvastatin Calcium", True),
    ("topiramate",   "Topiramate", True),
    ("tramadol",     "Acetaminophen| Tramadol Hydrochloride", True),
    ("metformin",    "Metformin Hydrochloride| Sitagliptin Phosphate Hydrate", True),
    ("amlodipine",   "Amlodipine Besylate| Atorvastatin Calcium Trihydrate", True),
    ("clopidogrel",  "Clopidogrel Bisulfate", True),
    ("warfarin",     "Warfarin Sodium", True),
    ("metformin + sitagliptin",
                     "Metformin Hydrochloride| Sitagliptin Phosphate Hydrate", True),
    ("rosuvastatin + ezetimibe",
                     "Ezetimibe| Rosuvastatin Calcium", True),
    # ── 매칭되면 안 되는 거짓(False) 케이스 — substring 함정 ──
    ("ephedrine",    "Pseudoephedrine Hydrochloride| Triprolidine Hydrochloride Hydrate", False),
    ("omeprazole",   "Esomeprazole Magnesium Trihydrate", False),
    ("ibuprofen",    "Dexibuprofen", False),
    ("codeine",      "Dihydrocodeine Tartrate", False),
    ("calcium chloride", "Calcium Pantothenate| Potassium Chloride| Magnesium Oxide", False),
    ("acetazolamide", "Topiramate", False),
    ("naproxen",     "Esomeprazole Magnesium Trihydrate", False),
]


def eval_matcher(matcher) -> dict:
    """gold set 에 대해 TP/FP/FN/TN, precision/recall/accuracy 계산."""
    tp = fp = fn = tn = 0
    wrong = []
    for inn, mat, expected in GOLD:
        pred = matcher(inn, mat)
        if pred and expected:       tp += 1
        elif pred and not expected: fp += 1; wrong.append(("FP", inn, mat))
        elif not pred and expected: fn += 1; wrong.append(("FN", inn, mat))
        else:                       tn += 1
    prec = tp / (tp + fp) if tp + fp else 0.0
    rec  = tp / (tp + fn) if tp + fn else 0.0
    acc  = (tp + tn) / len(GOLD)
    f1   = 2 * prec * rec / (prec + rec) if prec + rec else 0.0
    return dict(tp=tp, fp=fp, fn=fn, tn=tn,
                precision=prec, recall=rec, accuracy=acc, f1=f1, wrong=wrong)


# ─────────────── 전체 데이터셋 금기 쌍 재집계 ─────────────────────

def load_dataset_drugs() -> dict:
    raw = json.load(open(CLASS_MAP, encoding="utf-8"))
    return {code: {"name": v["name"], "material_en": v["material_en"]}
            for code, v in raw.items()}


def load_db_pairs() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    pairs = []
    for r in range(2, ws.max_row + 1):
        m1 = ws.cell(r, 2).value
        m2 = ws.cell(r, 3).value
        if m1 and m2:
            pairs.append({"m1": str(m1), "m2": str(m2),
                          "note": ws.cell(r, 4).value or "",
                          "info": ws.cell(r, 5).value or ""})
    return pairs


def detect_pairs(drugs: dict, db: list, matcher) -> set:
    """데이터셋에서 (약물코드1, 약물코드2) 금기 쌍 집합 반환."""
    found = set()
    for pair in db:
        a = [c for c, d in drugs.items() if matcher(pair["m1"], d["material_en"])]
        b = [c for c, d in drugs.items() if matcher(pair["m2"], d["material_en"])]
        if not a or not b:
            continue
        for ca in a:
            for cb in b:
                if ca != cb:
                    found.add(tuple(sorted((ca, cb))))
    return found


# ─────────────────────────── 실행 ────────────────────────────────

def main():
    print("=" * 62)
    print("  1) 라벨링 검증셋(gold set)으로 매칭 정확도 비교")
    print("=" * 62)
    print(f"  검증 케이스: {len(GOLD)}개 "
          f"(참 {sum(1 for *_ ,e in GOLD if e)} / 거짓 {sum(1 for *_ ,e in GOLD if not e)})\n")

    res_old = eval_matcher(match_substring)
    res_new = eval_matcher(match_token)

    hdr = f"  {'지표':<12}{'기존(substring)':>18}{'개선(token)':>16}"
    print(hdr); print("  " + "-" * 44)
    for key, label in [("precision", "Precision"), ("recall", "Recall"),
                       ("f1", "F1"), ("accuracy", "Accuracy")]:
        print(f"  {label:<12}{res_old[key]:>18.3f}{res_new[key]:>16.3f}")
    print(f"  {'FP(오경보)':<12}{res_old['fp']:>18d}{res_new['fp']:>16d}")
    print(f"  {'FN(놓침)':<12}{res_old['fn']:>18d}{res_new['fn']:>16d}")

    if res_old["wrong"]:
        print("\n  기존 방식이 틀린 케이스:")
        for kind, inn, mat in res_old["wrong"]:
            print(f"    [{kind}] INN={inn!r} vs {mat!r}")
    if res_new["wrong"]:
        print("\n  개선 방식이 틀린 케이스:")
        for kind, inn, mat in res_new["wrong"]:
            print(f"    [{kind}] INN={inn!r} vs {mat!r}")
    else:
        print("\n  개선 방식: 검증셋 전 케이스 정답 ✓")

    # ── 2) 전체 데이터셋 재집계 ──
    print("\n" + "=" * 62)
    print("  2) 전체 데이터셋(100클래스) 금기 쌍 재집계")
    print("=" * 62)
    drugs = load_dataset_drugs()
    db = load_db_pairs()
    old_pairs = detect_pairs(drugs, db, match_substring)
    new_pairs = detect_pairs(drugs, db, match_token)
    removed = old_pairs - new_pairs   # 개선으로 제거된(거짓 양성 추정) 쌍
    added   = new_pairs - old_pairs

    print(f"  기존(substring) 검출 쌍: {len(old_pairs)}개")
    print(f"  개선(token)    검출 쌍: {len(new_pairs)}개")
    print(f"  개선으로 제거된 쌍(거짓 양성 추정): {len(removed)}개")
    print(f"  개선으로 추가된 쌍              : {len(added)}개")

    if removed:
        print("\n  제거된 쌍 예시(기존의 오경보):")
        for ca, cb in list(removed)[:8]:
            print(f"    {drugs[ca]['name'][:24]} ↔ {drugs[cb]['name'][:24]}")

    # ── 저장 ──
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json.dump({
        "gold_set_size": len(GOLD),
        "substring": {k: res_old[k] for k in
                      ("precision", "recall", "f1", "accuracy", "fp", "fn")},
        "token":     {k: res_new[k] for k in
                      ("precision", "recall", "f1", "accuracy", "fp", "fn")},
        "dataset_pairs": {
            "substring": len(old_pairs),
            "token": len(new_pairs),
            "removed_false_positive": len(removed),
            "added": len(added),
        },
    }, open(OUT_DIR / "contraindication_validation.json", "w", encoding="utf-8"),
        ensure_ascii=False, indent=2)
    print(f"\n  결과 저장: {OUT_DIR / 'contraindication_validation.json'}")


if __name__ == "__main__":
    main()
