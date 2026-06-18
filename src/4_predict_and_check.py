"""
4_predict_and_check.py
======================
학습된 DrugCNN으로 이미지를 추론하고,
병용금기 DB를 조회해 위험 약물 조합을 경고합니다.

사용 예시:
    # 이미지 파일 여러 장을 직접 지정
    python src/4_predict_and_check.py --images path/to/drug1.png path/to/drug2.png

    # 원본 4-약물 복합 사진 1장 → bbox 크롭 후 추론
    python src/4_predict_and_check.py --photo path/to/combo_photo.png --json path/to/annotation.json

출력 예시:
    [결과] 인식된 약물:
      - 아질렉트정(라사길린메실산염)   (신뢰도 94.3%)
      - 울트라셋이알서방정             (신뢰도 87.1%)

    ⚠️  [경고] 병용금기 조합 발견!
      rasagiline ↔ tramadol
      → 세로토닌증후군 발생 위험 증가
"""

import argparse
import json
import re
import sys
from pathlib import Path

import torch
from torchvision import transforms
from PIL import Image

SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
CLASS_MAP_PATH   = PROJECT_ROOT / "data" / "class_map.json"
CONTRAIND_PATH   = PROJECT_ROOT / "data" / "contraindicated_drugs.xlsx"
DEFAULT_CKPT     = PROJECT_ROOT / "output" / "checkpoints" / "best_model.pth"


def load_class_map() -> dict:
    """
    class_map.json → {class_idx: {name, material_en, material_kr}}
    """
    raw = json.load(open(CLASS_MAP_PATH, encoding="utf-8"))
    return {
        info["class_idx"]: {
            "code":        code,
            "name":        info["name"],
            "material_en": info.get("material_en", ""),
            "material_kr": info.get("material_kr", ""),
        }
        for code, info in raw.items()
    }


def clean_ingredient(s: str) -> str:
    """괄호 제거 후 소문자 정규화"""
    return re.sub(r"\(.*?\)", "", str(s)).strip().lower()


def extract_ingredients(material_en: str) -> list[str]:
    """'|' 로 구분된 복합 성분 분리 + 정규화"""
    return [clean_ingredient(p) for p in material_en.split("|") if p.strip()]


def load_contraindicated_db() -> list[dict]:
    """병용금기 Excel → [{m1, m2, note, info}, ...]"""
    import openpyxl
    wb = openpyxl.load_workbook(CONTRAIND_PATH)
    ws = wb.active
    pairs = []
    for r in range(2, ws.max_row + 1):
        m1   = clean_ingredient(ws.cell(r, 2).value or "")
        m2   = clean_ingredient(ws.cell(r, 3).value or "")
        if m1 and m2:
            pairs.append({
                "m1":   m1,
                "m2":   m2,
                "note": ws.cell(r, 4).value or "",
                "info": ws.cell(r, 5).value or "",
            })
    return pairs


def ingredient_matches(excel_inn: str, drug_ingredients: list[str]) -> bool:
    """Excel INN이 약물 성분명의 substring인지 확인 (짧은 INN은 정확 일치)"""
    for ing in drug_ingredients:
        if len(excel_inn) <= 3:
            if excel_inn == ing:
                return True
        else:
            if excel_inn in ing:
                return True
    return False


def check_contraindications(
    drug_list: list[dict],
    db: list[dict],
) -> list[dict]:
    """
    인식된 약물 목록에서 병용금기 쌍 탐색.
    drug_list: [{name, material_en, ...}, ...]
    반환: [{drug_a, drug_b, m1, m2, info, note}, ...]
    """
    warnings = []

    drug_ingredients = [
        (d, extract_ingredients(d.get("material_en", "")))
        for d in drug_list
    ]

    for pair in db:
        m1, m2 = pair["m1"], pair["m2"]
        matched_a = [d for d, ings in drug_ingredients if ingredient_matches(m1, ings)]
        matched_b = [d for d, ings in drug_ingredients if ingredient_matches(m2, ings)]

        if matched_a and matched_b and set(id(d) for d in matched_a) != set(id(d) for d in matched_b):
            for da in matched_a:
                for db_ in matched_b:
                    warnings.append({
                        "drug_a": da["name"],
                        "drug_b": db_["name"],
                        "m1":     m1,
                        "m2":     m2,
                        "info":   pair["info"],
                        "note":   pair["note"],
                    })
    return warnings


def load_model(ckpt_path: Path, num_classes: int, device: torch.device):
    """저장된 체크포인트에서 모델 복원"""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "model_2", SCRIPT_DIR / "2_model.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    ckpt  = torch.load(ckpt_path, map_location=device)
    model = mod.DrugCNN(num_classes=num_classes).to(device)
    model.load_state_dict(ckpt["model_state"])
    model.eval()
    print(f"[추론] 모델 로드 완료: {ckpt_path}")
    print(f"       학습 epoch={ckpt.get('epoch')}, val_acc={ckpt.get('val_acc', 0):.4f}")
    return model


def preprocess_image(img_path: Path) -> torch.Tensor:
    """단일 이미지 → 정규화된 텐서 (1, 3, 128, 128)"""
    transform = transforms.Compose([
        transforms.Resize((128, 128)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])
    img = Image.open(img_path).convert("RGB")
    return transform(img).unsqueeze(0)


def crop_from_annotation(photo_path: Path, json_path: Path) -> list[tuple[Image.Image, dict]]:
    """
    원본 복합 사진 + COCO JSON → 개별 약물 크롭 이미지 목록
    반환: [(cropped_PIL_image, img_info), ...]
    """
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    photo = Image.open(photo_path).convert("RGB")
    W, H  = photo.size
    crops = []

    images      = data.get("images", [])
    annotations = data.get("annotations", [])

    for img_info, ann in zip(images, annotations):
        bbox = ann.get("bbox", [])
        if len(bbox) != 4:
            continue
        x, y, w, h = [int(v) for v in bbox]
        crop = photo.crop((max(0, x), max(0, y), min(W, x+w), min(H, y+h)))
        crops.append((crop, img_info))

    return crops


@torch.no_grad()
def predict_images(
    image_paths: list[Path],
    model,
    class_map: dict,
    device: torch.device,
    top_k: int = 3,
) -> list[dict]:
    """
    이미지 목록 → 예측 결과 목록
    반환: [{name, code, material_en, confidence, top_k: [...]}, ...]
    """
    results = []
    for img_path in image_paths:
        tensor = preprocess_image(img_path).to(device)
        probs  = torch.softmax(model(tensor), dim=1)[0]

        top_probs, top_indices = probs.topk(top_k)
        best_idx  = top_indices[0].item()
        best_prob = top_probs[0].item()

        drug_info = class_map.get(best_idx, {"name": "알 수 없음", "code": "?", "material_en": ""})
        results.append({
            "img_path":    img_path,
            "name":        drug_info["name"],
            "code":        drug_info.get("code", "?"),
            "material_en": drug_info.get("material_en", ""),
            "confidence":  best_prob,
            "top_k": [
                {
                    "name":        class_map.get(top_indices[i].item(), {}).get("name", "?"),
                    "confidence":  top_probs[i].item(),
                }
                for i in range(top_k)
            ],
        })
    return results


def print_results(predictions: list[dict], warnings: list[dict]) -> None:
    print("\n" + "=" * 60)
    print("  [결과] 인식된 약물")
    print("=" * 60)
    for pred in predictions:
        conf = pred["confidence"] * 100
        print(f"  ✅ {pred['name']:<40} (신뢰도 {conf:.1f}%)")
        for t in pred["top_k"][1:]:
            print(f"       후보: {t['name']:<35} ({t['confidence']*100:.1f}%)")

    if warnings:
        print("\n" + "⚠️  " * 15)
        print("  ⚠️  [경고] 병용금기 조합 발견!")
        print("⚠️  " * 15)
        for w in warnings:
            print(f"\n  {w['drug_a']}  ↔  {w['drug_b']}")
            print(f"  성분: {w['m1']}  ↔  {w['m2']}")
            if w["info"]:
                print(f"  위험: {w['info']}")
            if w["note"]:
                print(f"  조건: {w['note']}")
    else:
        print("\n  ✅ 병용금기 조합이 발견되지 않았습니다.")

    print("=" * 60 + "\n")


def main(args: argparse.Namespace) -> None:
    device    = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    class_map = load_class_map()
    db        = load_contraindicated_db()
    model     = load_model(Path(args.checkpoint), len(class_map), device)

    if args.images:
        image_paths = [Path(p) for p in args.images]
    elif args.photo and args.json:
        import tempfile, os
        crops = crop_from_annotation(Path(args.photo), Path(args.json))
        tmpdir = Path(tempfile.mkdtemp())
        image_paths = []
        for i, (crop_img, _) in enumerate(crops):
            tp = tmpdir / f"crop_{i}.png"
            crop_img.resize((128, 128), Image.BILINEAR).save(tp)
            image_paths.append(tp)
        print(f"[추론] 복합 사진에서 {len(image_paths)}개 크롭 추출")
    else:
        print("[오류] --images 또는 --photo + --json 를 지정해주세요.")
        sys.exit(1)

    predictions = predict_images(image_paths, model, class_map, device)
    warnings = check_contraindications(predictions, db)
    print_results(predictions, warnings)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="DrugCNN Inference + Contraindication Check")
    p.add_argument("--images",     nargs="+", help="단일 약물 이미지 경로 목록")
    p.add_argument("--photo",      help="원본 복합 사진 경로")
    p.add_argument("--json",       help="복합 사진의 COCO 어노테이션 JSON 경로")
    p.add_argument("--checkpoint", default=str(DEFAULT_CKPT), help="모델 체크포인트 경로")
    main(p.parse_args())
